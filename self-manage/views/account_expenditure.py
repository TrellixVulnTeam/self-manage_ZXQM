# -*- coding: utf-8 -*-
import sqlite3

import wx
from .generic_bitmap_button import GenericBitmapButton
import wx.adv
import time
from .my_Validator import MyNumberValidator
from openpyxl import load_workbook
import wx.lib.buttons as buttons
import images
from wx import NewId
from views.Dialogs import NotExsit,AddSuccess,WriteFail,IncompleteData

expend = [u'教育',u'餐饮',u'理财',u'日用',u'零食',u'交通',u'服饰美容',u'数码',u'住房',u'医疗']
ID_00 = NewId()
ID_01 = NewId()
ID_02 = NewId()
ID_03 = NewId()
ID_04 = NewId()
ID_05 = NewId()
ID_06 = NewId()
ID_07 = NewId()
ID_08 = NewId()
ID_09 = NewId()

class Expenditure(wx.Panel):
    def __init__(self,parent):
        super().__init__(parent)

        self.n_time = time.strftime('%Y-%m-%d', time.localtime())
        self.n_kind = expend[0]    #默认消费种类值
        self.n_acount = 0
        self.n_remark = '无'         #消费备注
        self.initUi()
        self.Center()

    def initUi(self):
        self.SetBackgroundColour("white")

        box1 = wx.GridSizer(4,5,5,10)

        #生成按钮用来接受用户的输入
        self.t_0= wx.StaticText(self,label = expend[0])
        self.bt_0 = buttons.GenBitmapToggleButton(self, ID_00, None)
        bmp = images._10.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_0.SetBitmapLabel(bmp)
        bmp = images._100.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_0.SetBitmapSelected(bmp)
        self.bt_0.SetToggle(False)
        self.bt_0.SetInitialSize()
        self.bt_0.Bind(wx.EVT_BUTTON,self.KindSelect)

        self.t_1 = wx.StaticText(self, label=expend[1])
        self.bt_1 = buttons.GenBitmapToggleButton(self, ID_01, None)
        bmp = images._11.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_1.SetBitmapLabel(bmp)
        bmp = images._110.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_1.SetBitmapSelected(bmp)
        self.bt_1.SetToggle(False)
        self.bt_1.SetInitialSize()
        self.bt_1.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_2 = wx.StaticText(self, label=expend[2])
        self.bt_2 = buttons.GenBitmapToggleButton(self, ID_02, None)
        bmp = images._12.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_2.SetBitmapLabel(bmp)
        bmp = images._120.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_2.SetBitmapSelected(bmp)
        self.bt_2.SetToggle(False)
        self.bt_2.SetInitialSize()
        self.bt_2.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_3 = wx.StaticText(self, label=expend[3])
        self.bt_3 = buttons.GenBitmapToggleButton(self, ID_03, None)
        bmp = images._13.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_3.SetBitmapLabel(bmp)
        bmp = images._130.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_3.SetBitmapSelected(bmp)
        self.bt_3.SetToggle(False)
        self.bt_3.SetInitialSize()
        self.bt_3.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_4 = wx.StaticText(self, label=expend[4])
        self.bt_4 = buttons.GenBitmapToggleButton(self, ID_04, None)
        bmp = images._14.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_4.SetBitmapLabel(bmp)
        bmp = images._140.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_4.SetBitmapSelected(bmp)
        self.bt_4.SetToggle(False)
        self.bt_4.SetInitialSize()
        self.bt_4.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_5 = wx.StaticText(self, label=expend[5])
        self.bt_5 = buttons.GenBitmapToggleButton(self, ID_05, None)
        bmp = images._15.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_5.SetBitmapLabel(bmp)
        bmp = images._150.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_5.SetBitmapSelected(bmp)
        self.bt_5.SetToggle(False)
        self.bt_5.SetInitialSize()
        self.bt_5.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_6 = wx.StaticText(self, label=expend[6])
        self.bt_6 = buttons.GenBitmapToggleButton(self, ID_06, None)
        bmp = images._16.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_6.SetBitmapLabel(bmp)
        bmp = images._160.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_6.SetBitmapSelected(bmp)
        self.bt_6.SetToggle(False)
        self.bt_6.SetInitialSize()
        self.bt_6.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_7 = wx.StaticText(self, label=expend[7])
        self.bt_7 = buttons.GenBitmapToggleButton(self, ID_07, None)
        bmp = images._17.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_7.SetBitmapLabel(bmp)
        bmp = images._170.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_7.SetBitmapSelected(bmp)
        self.bt_7.SetToggle(False)
        self.bt_7.SetInitialSize()
        self.bt_7.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_8 = wx.StaticText(self, label=expend[8])
        self.bt_8 = buttons.GenBitmapToggleButton(self, ID_08, None)
        bmp = images._18.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_8.SetBitmapLabel(bmp)
        bmp = images._180.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_8.SetBitmapSelected(bmp)
        self.bt_8.SetToggle(False)
        self.bt_8.SetInitialSize()
        self.bt_8.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_9 = wx.StaticText(self, label=expend[9])
        self.bt_9 = buttons.GenBitmapToggleButton(self, ID_09, None)
        bmp = images._19.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_9.SetBitmapLabel(bmp)
        bmp = images._190.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_9.SetBitmapSelected(bmp)
        self.bt_9.SetToggle(False)
        self.bt_9.SetInitialSize()
        self.bt_9.Bind(wx.EVT_BUTTON, self.KindSelect)

        box1.AddMany([(self.bt_0,0,wx.EXPAND),(self.bt_1,0,wx.EXPAND),
                      (self.bt_2,0,wx.EXPAND),(self.bt_3,0,wx.EXPAND),
                      (self.bt_4,0,wx.EXPAND),
                      (self.t_0, 0, wx.LEFT),
                      (self.t_1, 0, wx.LEFT),(self.t_2, 0, wx.LEFT),
                      (self.t_3, 0, wx.LEFT),(self.t_4, 0, wx.LEFT),

                      (self.bt_5,0,wx.EXPAND),(self.bt_6,0,wx.EXPAND),
                      (self.bt_7,0,wx.EXPAND),(self.bt_8,0,wx.EXPAND),
                      (self.bt_9,0,wx.EXPAND),
                      (self.t_5, 0, wx.EXPAND),(self.t_6, 0, wx.EXPAND),
                      (self.t_7, 0, wx.EXPAND),
                      (self.t_8, 0, wx.EXPAND), (self.t_9, 0, wx.EXPAND),
                      ])

        box2 = wx.BoxSizer(wx.HORIZONTAL)

        #消费时间
        self.time1 = wx.adv.DatePickerCtrl(self,style = wx.adv.DP_DROPDOWN
                                      | wx.adv.DP_SHOWCENTURY
                                      | wx.adv.DP_ALLOWNONE)

        self.Bind(wx.adv.EVT_DATE_CHANGED,self.OnDateSelect,self.time1)

        self.t_acount = wx.TextCtrl(self, -1,"金额",validator = MyNumberValidator())
        self.Bind(wx.EVT_TEXT,self.EvtText1,self.t_acount)

        box2.Add(self.time1,flag=wx.LEFT,border=10)
        box2.Add(self.t_acount,flag = wx.RIGHT|wx.EXPAND,border =10)

        self.t_remarks = wx.TextCtrl(self,-1, "备注",size =(300,150))
        self.Bind(wx.EVT_TEXT, self.EvtText2, self.t_remarks)

        self.bt_commit = GenericBitmapButton(self, 'commit', -1)
        self.bt_commit.SetToolTip('提交')
        self.Bind(wx.EVT_BUTTON,self.Commit,self.bt_commit)

        box3 = wx.BoxSizer(wx.VERTICAL)

        box3.AddSpacer(10)
        box3.Add(box1,flag=wx.ALIGN_CENTER)
        box3.Add(box2,flag = wx.ALIGN_CENTER)
        box3.Add(self.t_remarks,flag=wx.ALIGN_CENTER|wx.LEFT|wx.RIGHT,border=20)
        box4= wx.BoxSizer(wx.HORIZONTAL)
        box4.AddSpacer(500)
        box4.Add(self.bt_commit,flag = wx.RIGHT,border = 10)
        box3.Add(box4,flag =wx.ALIGN_CENTER,border = 10)

        self.SetSizer(box3)

    def EvtText1(self,e):
        self.n_acount= str(e.GetString())

    def EvtText2(self,e):
        self.n_remark = e.GetString()

    def OnDateSelect(self,e):
        self.n_time = str(e.GetDate())[0:10]

    def KindSelect(self,e):

        #先将按钮设置全部设置为未选中
        self.bt_0.SetToggle(False)
        self.bt_1.SetToggle(False)
        self.bt_2.SetToggle(False)
        self.bt_3.SetToggle(False)
        self.bt_4.SetToggle(False)
        self.bt_5.SetToggle(False)
        self.bt_6.SetToggle(False)
        self.bt_7.SetToggle(False)
        self.bt_8.SetToggle(False)
        self.bt_9.SetToggle(False)
        if e.GetId() == ID_00:
            self.bt_0.SetToggle(True)
            self.n_kind = expend[0]
        if e.GetId() == ID_01:
            self.bt_1.SetToggle(True)
            self.n_kind =expend[1]
        if e.GetId() == ID_02:
            self.bt_2.SetToggle(True)
            self.n_kind =expend[2]
        if e.GetId() == ID_03:
            self.bt_3.SetToggle(True)
            self.n_kind =expend[3]
        if e.GetId() == ID_04:
            self.bt_4.SetToggle(True)
            self.n_kind =expend[4]
        if e.GetId() == ID_05:
            self.bt_5.SetToggle(True)
            self.n_kind =expend[5]
        if e.GetId() == ID_06:
            self.bt_6.SetToggle(True)
            self.n_kind =expend[6]
        if e.GetId() == ID_07:
            self.bt_7.SetToggle(True)
            self.n_kind =expend[7]
        if e.GetId() == ID_08:
            self.bt_8.SetToggle(True)
            self.n_kind =expend[8]
        if e.GetId() == ID_09:
            self.bt_9.SetToggle(True)
            self.n_kind =expend[9]


    def Commit(self,e):


        self.db_write()

    def db_write(self):
        # 使用该语句如果不存在该数据库则自动生成一个数据库

        if self.n_kind and self.n_time and self.n_acount:
            try:
                conn = sqlite3.connect('my_record.db')

                c = conn.cursor()

                # 消费记录
                c.execute('''CREATE TABLE IF NOT EXISTS EXPENDITURES
                       (ID INT PRIMARY KEY     ,
                       XTime          TEXT    NOT NULL,
                       KIND            TEXT     NOT NULL,
                       ACCOUNT        DOUBLE   NOT NULL,
                       REMARK         TEXT     NOT NULL);''')


                c.execute("INSERT INTO EXPENDITURES(XTime,KIND,ACCOUNT,REMARK ) VALUES (?,?,?,?)",
                          (self.n_time,self.n_kind,self.n_acount,self.n_remark))

                conn.commit()
                conn.close()

                dlg = AddSuccess(None, -1)
                dlg.ShowModal()
                dlg.Destroy()

                # 设置禁用时间
                self.bt_commit.Enable(False)
                time.sleep(1)
                self.bt_commit.Enable(True)

            except:
                dlg = WriteFail(None, -1)
                dlg.ShowModal()
                dlg.Destroy()



                # 设置禁用时间
                self.bt_commit.Enable(False)
                time.sleep(1)
                self.bt_commit.Enable(True)

        else:
            dlg = IncompleteData(None, -1)
            dlg.ShowModal()
            dlg.Destroy()

            #设置禁用时间
            self.bt_commit.Enable(False)
            time.sleep(1)
            self.bt_commit.Enable(True)




    def excel_write(self):

        #写入excel表格中
        try:
            wb=load_workbook('./data/account.xlsx')
            ws=wb['支出']
            b=int(ws['A1'].value)    #excel 单元格第一格用来记录上次写入的位置
            a=ws.cell(row = b,column = 1)

        except:
            dlg = NotExsit(None, -1)
            dlg.ShowModal()
            dlg.Destroy()

        #再一次判断b值是否设置有错，防止出现删除消费记录但未更新的情况
        while a.value:
            b+=1
            a=ws['A'+str(b)]
            if b>=1500:
                break


        if self.n_kind and self.n_time and self.n_acount:

            try:
                g = str(b)
                ws['A'+g]=self.n_time
                ws['B'+g]=self.n_kind
                ws['C'+g]=self.n_acount
                ws['D'+g]=self.n_remark
                ws['A1'] = b+1

                wb.save('./data/account.xlsx')
                dlg = AddSuccess(None, -1)
                dlg.ShowModal()
                dlg.Destroy()


            except:
                dlg = WriteFail(None, -1)
                dlg.ShowModal()
                dlg.Destroy()


        else:
            dlg = IncompleteData(None,-1)
            dlg.ShowModal()
            dlg.Destroy()

        self.bt_commit.Enable(False)
        time.sleep(1)
        self.bt_commit.Enable(True)
















